"""
GNMA Multifamily Data Downloader & S-Curve Builder
===================================================
Run this locally or in a web IDE (Replit, Colab, Codespaces).

Authentication strategy:
  1. Try requests + BeautifulSoup (fast, no browser needed)
  2. If that fails, auto-fallback to Playwright headless Chromium
     (auto-installs playwright + chromium if missing)

Requirements (base):
  pip install requests beautifulsoup4 pandas openpyxl lxml

Playwright is auto-installed only if the requests method fails.

Usage:
  python gnma_mf_downloader.py
  python gnma_mf_downloader.py --email you@email.com --answer "Red" --months 6
  python gnma_mf_downloader.py --skip-download --data-dir ./my_files
  python gnma_mf_downloader.py --browser          # force Playwright mode
  python gnma_mf_downloader.py --headed            # visible browser (debug)
"""

import argparse
import glob
import os
import re
import subprocess
import sys
import time
import zipfile
from collections import defaultdict
from datetime import datetime, timedelta

try:
    import requests
    from bs4 import BeautifulSoup
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install requests beautifulsoup4 pandas openpyxl lxml")
    sys.exit(1)


# ─── CONFIGURATION ────────────────────────────────────────
BASE_URL = "https://www.ginniemae.gov"
BULK_URL = "https://bulk.ginniemae.gov"
PROFILE_URL = f"{BASE_URL}/Pages/profile.aspx"
DOWNLOAD_PAGE = f"{BASE_URL}/data_and_reports/disclosure_data/Pages/datadownload_bulk.aspx"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "gnma_mf_data")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "gnma_mf_scurve_dataset.xlsx")

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36")

POOL_TYPE_NAMES = {
    'PL': 'Project Loan', 'PN': 'Non-Level PL', 'LM': 'Mature/Modified',
    'LS': 'Small PL', 'RX': 'Mark-to-Market',
    'CL': 'Construction (Same)', 'CS': 'Construction (Diff)',
}
SCURVE_POOL_TYPES = {'PL', 'PN', 'LM', 'LS', 'RX'}

P = {
    'cusip': 0, 'pool_number': 1, 'pool_indicator': 2, 'pool_type': 3,
    'security_rate': 4, 'issue_date': 5, 'maturity_date': 6, 'orig_agg_amount': 7,
    'issuer_number': 8, 'issuer_name': 9,
    'pool_upb': 16, 'num_loans': 17,
    'num_30dq': 18, 'upb_30dq': 19, 'pct_30dq': 20,
    'num_60dq': 21, 'upb_60dq': 22, 'pct_60dq': 23,
    'num_90dq': 24, 'upb_90dq': 25, 'pct_90dq': 26,
    'security_rpb': 27, 'rpb_factor': 28,
    'proj_loan_sec_rate': 29, 'est_mtg_amount': 30,
}
L = {
    'disclosure_seq': 31, 'case_number': 32, 'agency_type': 33, 'loan_type': 34,
    'loan_term': 35, 'first_pay_date': 36, 'loan_maturity_date': 37, 'loan_rate': 38,
    'modified_ind': 39, 'non_level_ind': 40, 'mature_loan_flag': 41,
    'origination_date': 42, 'initial_endorsement': 43, 'final_endorsement': 44,
    'lockout_term': 45, 'lockout_end_date': 46,
    'prepay_premium_period': 47, 'prepay_end_date': 48,
    'interest_approval_date': 49, 'prepay_penalty_flag': 50,
    'orig_prin_bal': 51, 'upb_at_issuance': 52, 'upb': 53,
    'draw_number': 54, 'approved_draw_amt': 55,
    'months_dq': 56, 'liquidation_flag': 57, 'removal_reason': 58,
    'seller_issuer_id': 59,
    'property_name': 60, 'property_street': 61, 'property_city': 62,
    'property_state': 63, 'property_zip': 64, 'msa': 65,
    'num_units': 66, 'pi_amount': 67,
    'prepay_desc': 68, 'non_level_desc': 69,
    'fha_program_code': 70, 'insurance_type': 71,
    'as_of_date': 72, 'green_status': 73, 'affordable_status': 74,
}


# ═══════════════════════════════════════════════════════════
#  METHOD 1: REQUESTS-BASED AUTH (fast, no browser)
# ═══════════════════════════════════════════════════════════

class RequestsDownloader:

    def __init__(self, email, answer):
        self.email = email
        self.answer = answer
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": UA})
        self.authenticated = False

    def _aspnet_fields(self, html):
        soup = BeautifulSoup(html, "lxml")
        fields = {}
        for n in ("__VIEWSTATE", "__VIEWSTATEGENERATOR",
                  "__EVENTVALIDATION", "__REQUESTDIGEST"):
            tag = soup.find("input", {"name": n})
            if tag:
                fields[n] = tag.get("value", "")
        return fields, soup

    def _find(self, soup, keywords, types=("text", "email", "password")):
        for inp in soup.find_all("input", {"type": list(types)}):
            nm = (inp.get("name") or "").lower()
            iid = (inp.get("id") or "").lower()
            if any(k in nm or k in iid for k in keywords):
                return inp.get("name")
        return None

    def authenticate(self):
        print("\n[requests] Attempting authentication...")
        try:
            resp = self.session.get(DOWNLOAD_PAGE, timeout=30, allow_redirects=True)
            resp.raise_for_status()
        except requests.RequestException as e:
            print(f"  x Failed: {e}")
            return False

        if "datadownload_bulk" in resp.url.lower() and "profile" not in resp.url.lower():
            print("  ok No auth needed")
            self.authenticated = True
            return True

        fields, soup = self._aspnet_fields(resp.text)
        ef = self._find(soup, ["email", "mail"])
        af = self._find(soup, ["answer", "security", "question"])
        sf = self._find(soup, ["submit", "btn"], types=("submit", "button", "image"))

        if not ef or not af:
            print("  x Could not find form fields (page likely requires JavaScript)")
            return False

        print(f"  Found: email={ef}, answer={af}")
        post = {**fields, ef: self.email, af: self.answer}
        if sf:
            post[sf] = "Submit"

        try:
            resp = self.session.post(resp.url, data=post, timeout=30, allow_redirects=True)
        except requests.RequestException as e:
            print(f"  x POST failed: {e}")
            return False

        # Verify
        try:
            check = self.session.get(DOWNLOAD_PAGE, timeout=30, allow_redirects=True)
            if "profile" not in check.url.lower():
                print("  ok Authentication successful!")
                self.authenticated = True
                return True
        except:
            pass

        if "data_bulk" in resp.text.lower() or "mfplmon3" in resp.text.lower():
            print("  ok Authentication successful!")
            self.authenticated = True
            return True

        print("  x Auth failed (page requires JavaScript)")
        return False

    def download_file(self, url, filename, dest_dir):
        dest = os.path.join(dest_dir, filename)
        if os.path.exists(dest) and os.path.getsize(dest) > 1000:
            print(f"  ok {filename} (cached, {os.path.getsize(dest)//1024} KB)")
            return dest
        try:
            resp = self.session.get(url, timeout=120, stream=True, allow_redirects=True)
            if "profile.aspx" in resp.url.lower():
                print(f"  x {filename} — session expired")
                return None
            ct = resp.headers.get("Content-Type", "")
            if "text/html" in ct and "profile" in resp.text[:500].lower():
                print(f"  x {filename} — auth wall")
                return None
            resp.raise_for_status()
            total = int(resp.headers.get("Content-Length", 0))
            dl = 0
            with open(dest, "wb") as f:
                for chunk in resp.iter_content(65536):
                    f.write(chunk)
                    dl += len(chunk)
                    if total:
                        print(f"\r  > {filename} {dl//1024}/{total//1024} KB ({dl*100//total}%)",
                              end="", flush=True)
            sz = os.path.getsize(dest)
            print(f"\r  ok {filename} — {sz//1024} KB" + " "*30)
            if sz < 1000:
                with open(dest, 'r', errors='replace') as f:
                    if 'html' in f.read(300).lower():
                        os.remove(dest)
                        return None
            return dest
        except requests.RequestException as e:
            print(f"  x {filename} — {e}")
            return None


# ═══════════════════════════════════════════════════════════
#  METHOD 2: PLAYWRIGHT BROWSER (auto-fallback)
# ═══════════════════════════════════════════════════════════

def ensure_playwright():
    try:
        from playwright.sync_api import sync_playwright
        return True
    except ImportError:
        print("\n[playwright] Installing playwright + chromium...")
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "playwright"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            subprocess.check_call(
                [sys.executable, "-m", "playwright", "install", "chromium"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            try:
                subprocess.check_call(
                    [sys.executable, "-m", "playwright", "install-deps", "chromium"],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except:
                pass
            from playwright.sync_api import sync_playwright
            print("[playwright] ok Installed")
            return True
        except Exception as e:
            print(f"[playwright] x Install failed: {e}")
            print("  Manual: pip install playwright && python -m playwright install chromium")
            return False


class PlaywrightDownloader:

    def __init__(self, email, answer, headless=True):
        self.email = email
        self.answer = answer
        self.headless = headless
        self.authenticated = False
        self._pw = None
        self._browser = None
        self._ctx = None
        self._page = None

    def _start(self):
        from playwright.sync_api import sync_playwright
        self._pw = sync_playwright().start()
        self._browser = self._pw.chromium.launch(headless=self.headless)
        self._ctx = self._browser.new_context(
            user_agent=UA, accept_downloads=True,
            viewport={"width": 1280, "height": 900})
        self._page = self._ctx.new_page()

    def close(self):
        try:
            if self._ctx: self._ctx.close()
            if self._browser: self._browser.close()
            if self._pw: self._pw.stop()
        except:
            pass

    def authenticate(self):
        print("\n[playwright] Launching headless Chromium...")
        try:
            self._start()
        except Exception as e:
            print(f"  x Browser launch failed: {e}")
            return False

        page = self._page
        try:
            print("[playwright] Navigating to GNMA...")
            page.goto(DOWNLOAD_PAGE, wait_until="networkidle", timeout=60000)
            time.sleep(2)
            url = page.url.lower()
            print(f"  URL: {page.url}")

            if "profile" not in url:
                print("  ok Already authenticated")
                self.authenticated = True
                return True

            print("[playwright] On profile page — filling credentials...")

            # ── Find and fill email ──
            filled_email = False
            for sel in ['input[name*="Email" i]', 'input[id*="Email" i]',
                        'input[type="email"]', 'input[name*="txtEmail"]',
                        'input[placeholder*="email" i]']:
                try:
                    el = page.query_selector(sel)
                    if el and el.is_visible():
                        el.fill(self.email)
                        filled_email = True
                        print(f"  ok Email: {sel}")
                        break
                except:
                    continue

            # ── Find and fill answer ──
            filled_answer = False
            for sel in ['input[name*="Answer" i]', 'input[id*="Answer" i]',
                        'input[name*="Security" i]', 'input[name*="txtAnswer"]',
                        'input[placeholder*="answer" i]']:
                try:
                    el = page.query_selector(sel)
                    if el and el.is_visible():
                        el.fill(self.answer)
                        filled_answer = True
                        print(f"  ok Answer: {sel}")
                        break
                except:
                    continue

            # ── Positional fallback ──
            if not filled_email or not filled_answer:
                print("  Trying positional fallback...")
                inputs = page.query_selector_all(
                    'input[type="text"], input[type="email"], input[type="password"]')
                vis = [i for i in inputs if i.is_visible()]
                print(f"  {len(vis)} visible inputs found")

                for inp in vis:
                    nm = (inp.get_attribute("name") or "").lower()
                    iid = (inp.get_attribute("id") or "").lower()
                    ph = (inp.get_attribute("placeholder") or "").lower()
                    sig = f"{nm} {iid} {ph}"
                    print(f"    name={nm!r} id={iid!r} ph={ph!r}")

                    if not filled_email and any(k in sig for k in ("email", "mail")):
                        inp.fill(self.email); filled_email = True
                        print(f"    -> email")
                    elif not filled_answer and any(k in sig for k in ("answer", "security", "color", "question")):
                        inp.fill(self.answer); filled_answer = True
                        print(f"    -> answer")

                # Last resort: 2 inputs = email, answer by position
                if (not filled_email or not filled_answer) and len(vis) == 2:
                    vis[0].fill(self.email); vis[1].fill(self.answer)
                    filled_email = filled_answer = True
                    print("  ok Positional fill (2 inputs)")

            if not filled_email or not filled_answer:
                ss = os.path.join(SCRIPT_DIR, "gnma_debug.png")
                page.screenshot(path=ss, full_page=True)
                print(f"  x Could not fill form. Screenshot: {ss}")
                return False

            # ── Submit ──
            print("[playwright] Submitting...")
            clicked = False
            for sel in ['input[type="submit"]', 'button[type="submit"]',
                        'input[value*="Submit" i]', 'input[value*="Go" i]',
                        'button:has-text("Submit")', 'input[name*="btnSubmit"]']:
                try:
                    btn = page.query_selector(sel)
                    if btn and btn.is_visible():
                        btn.click(); clicked = True
                        print(f"  ok Click: {sel}")
                        break
                except:
                    continue
            if not clicked:
                page.keyboard.press("Enter")
                print("  ok Pressed Enter")

            try:
                page.wait_for_load_state("networkidle", timeout=30000)
            except:
                time.sleep(5)
            time.sleep(3)

            # ── Verify ──
            url = page.url.lower()
            if "profile" in url:
                # Try navigating again
                page.goto(DOWNLOAD_PAGE, wait_until="networkidle", timeout=30000)
                time.sleep(3)
                url = page.url.lower()

            if "profile" not in url or "data_bulk" in page.content().lower():
                print("  ok Authentication successful!")
                self.authenticated = True
                return True

            ss = os.path.join(SCRIPT_DIR, "gnma_debug.png")
            page.screenshot(path=ss, full_page=True)
            print(f"  x Still on profile. Screenshot: {ss}")
            return False

        except Exception as e:
            print(f"  x Error: {e}")
            try:
                page.screenshot(path=os.path.join(SCRIPT_DIR, "gnma_debug.png"), full_page=True)
            except:
                pass
            return False

    def download_file(self, url, filename, dest_dir):
        dest = os.path.join(dest_dir, filename)
        if os.path.exists(dest) and os.path.getsize(dest) > 1000:
            print(f"  ok {filename} (cached, {os.path.getsize(dest)//1024} KB)")
            return dest

        page = self._page
        try:
            # Method A: Playwright download interception
            with page.expect_download(timeout=120000) as dl_info:
                page.goto(url, timeout=60000)
            download = dl_info.value
            download.save_as(dest)
            sz = os.path.getsize(dest)
            print(f"  ok {filename} — {sz//1024} KB")
            if sz < 1000:
                with open(dest, 'r', errors='replace') as f:
                    if 'html' in f.read(300).lower():
                        os.remove(dest)
                        print(f"    x Was HTML, removed")
                        return None
            return dest
        except Exception:
            pass

        try:
            # Method B: Direct response capture
            resp = page.goto(url, wait_until="commit", timeout=60000)
            if resp and resp.status == 200:
                body = resp.body()
                if len(body) > 1000 and body[:2] == b'PK':
                    with open(dest, 'wb') as f:
                        f.write(body)
                    print(f"  ok {filename} — {len(body)//1024} KB (direct)")
                    return dest
        except:
            pass

        try:
            # Method C: Use cookies from Playwright in a requests session
            cookies = self._ctx.cookies()
            s = requests.Session()
            for c in cookies:
                s.cookies.set(c['name'], c['value'], domain=c.get('domain', ''))
            s.headers.update({"User-Agent": UA})
            resp = s.get(url, timeout=120, stream=True, allow_redirects=True)
            if "profile" not in resp.url.lower() and resp.status_code == 200:
                with open(dest, 'wb') as f:
                    for chunk in resp.iter_content(65536):
                        f.write(chunk)
                sz = os.path.getsize(dest)
                if sz > 1000:
                    print(f"  ok {filename} — {sz//1024} KB (cookie transfer)")
                    return dest
                else:
                    os.remove(dest)
        except:
            pass

        print(f"  x {filename} — all download methods failed")
        return None


# ═══════════════════════════════════════════════════════════
#  DOWNLOAD ORCHESTRATOR
# ═══════════════════════════════════════════════════════════

def get_file_list(months=6):
    now = datetime.now()
    files = []
    for i in range(months + 2):
        dt = now - timedelta(days=30 * (i + 1))
        period = dt.strftime("%Y%m")
        fn = f"mfplmon3_{period}.zip"
        url = f"{BULK_URL}/protectedfiledownload.aspx?dlfile=data_bulk/{fn}"
        files.append({"period": period, "filename": fn, "url": url})
    return files[:months]


def download_all(dl, months=6):
    os.makedirs(DATA_DIR, exist_ok=True)
    files = get_file_list(months)
    print(f"\nDownloading {len(files)} files to {DATA_DIR}\n")
    ok = 0
    for f in files:
        if dl.download_file(f["url"], f["filename"], DATA_DIR):
            ok += 1
        time.sleep(1)
    print(f"\n  {ok}/{len(files)} files downloaded")
    return ok


# ═══════════════════════════════════════════════════════════
#  FILE PARSING
# ═══════════════════════════════════════════════════════════

def sf(s):
    try: return float(s.strip()) if s and s.strip() else np.nan
    except: return np.nan

def si(s):
    try: return int(float(s.strip())) if s and s.strip() else 0
    except: return 0

def pd_(s):
    s = (s or "").strip()
    if not s or len(s) < 6: return pd.NaT
    try:
        if len(s) == 8: return pd.Timestamp(f"{s[:4]}-{s[4:6]}-{s[6:8]}")
        if len(s) == 6: return pd.Timestamp(f"{s[:4]}-{s[4:6]}-01")
    except: pass
    return pd.NaT

def mb(d1, d2):
    if pd.isna(d1) or pd.isna(d2): return np.nan
    return (d2.year - d1.year) * 12 + (d2.month - d1.month)


def read_mfplmon3(filepath):
    text = None
    if filepath.endswith('.zip'):
        try:
            with zipfile.ZipFile(filepath, 'r') as zf:
                for name in zf.namelist():
                    if name.endswith('.txt') or 'mfplmon' in name.lower():
                        with zf.open(name) as f:
                            text = f.read().decode('utf-8', errors='replace')
                        break
                if text is None and zf.namelist():
                    with zf.open(zf.namelist()[0]) as f:
                        text = f.read().decode('utf-8', errors='replace')
        except zipfile.BadZipFile:
            return []
    else:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read()
    if not text: return []

    records = []
    for line in text.split('\n'):
        flds = line.rstrip('\r').split('|')
        if len(flds) < 32: continue
        cusip = flds[0].strip()
        if not cusip or len(cusip) != 9 or not cusip[0].isdigit(): continue
        pt = flds[P['pool_type']].strip()

        def g(idx):
            return flds[idx].strip() if len(flds) > idx else ''

        rec = {
            'pool_cusip': cusip, 'pool_number': g(P['pool_number']),
            'pool_type': pt, 'pool_type_name': POOL_TYPE_NAMES.get(pt, pt),
            'security_rate': sf(g(P['security_rate'])),
            'issue_date': g(P['issue_date']),
            'pool_maturity_date': g(P['maturity_date']),
            'orig_agg_amount': sf(g(P['orig_agg_amount'])),
            'issuer_number': g(P['issuer_number']), 'issuer_name': g(P['issuer_name']),
            'pool_upb': sf(g(P['pool_upb'])),
            'security_rpb': sf(g(P['security_rpb'])),
            'rpb_factor': sf(g(P['rpb_factor'])),
        }

        if len(flds) > L['case_number']:
            case = g(L['case_number'])
            rec.update({
                'case_number': case, 'loan_id': cusip + '_' + case,
                'agency_type': g(L['agency_type']), 'loan_type': g(L['loan_type']),
                'loan_term': si(g(L['loan_term'])),
                'first_pay_date': g(L['first_pay_date']),
                'loan_maturity_date': g(L['loan_maturity_date']),
                'loan_rate': sf(g(L['loan_rate'])),
                'modified_ind': g(L['modified_ind']),
                'non_level_ind': g(L['non_level_ind']),
                'mature_loan_flag': g(L['mature_loan_flag']),
                'origination_date': g(L['origination_date']),
                'lockout_term_yrs': si(g(L['lockout_term'])),
                'lockout_end_date': g(L['lockout_end_date']),
                'prepay_premium_period_yrs': si(g(L['prepay_premium_period'])),
                'prepay_end_date': g(L['prepay_end_date']),
                'prepay_penalty_flag': g(L['prepay_penalty_flag']),
                'orig_prin_bal': sf(g(L['orig_prin_bal'])),
                'upb_at_issuance': sf(g(L['upb_at_issuance'])),
                'upb': sf(g(L['upb'])),
                'months_dq': si(g(L['months_dq'])),
                'liquidation_flag': g(L['liquidation_flag']),
                'removal_reason': g(L['removal_reason']),
                'property_name': g(L['property_name']),
                'property_city': g(L['property_city']),
                'property_state': g(L['property_state']),
                'msa': g(L['msa']),
                'num_units': si(g(L['num_units'])),
                'pi_amount': sf(g(L['pi_amount'])),
                'prepay_desc': g(L['prepay_desc']),
                'fha_program_code': g(L['fha_program_code']),
                'insurance_type': g(L['insurance_type']),
                'as_of_date': g(L['as_of_date']),
                'green_status': g(L['green_status']),
                'affordable_status': g(L['affordable_status']),
            })
        else:
            rec.update({'loan_id': cusip + '_POOL', 'case_number': '',
                        'loan_rate': np.nan, 'upb': np.nan})
        records.append(rec)
    return records


# ═══════════════════════════════════════════════════════════
#  PANEL & S-CURVE
# ═══════════════════════════════════════════════════════════

def build_panel(monthly_data):
    periods = sorted(monthly_data.keys())
    print(f"\nBuilding panel: {len(periods)} periods")
    pmap = {p: {r['loan_id']: r for r in recs} for p, recs in monthly_data.items()}
    rows = []
    for i, period in enumerate(periods):
        nxt = periods[i+1] if i+1 < len(periods) else None
        nl = pmap.get(nxt, {}) if nxt else {}
        od = pd_(period)
        for lid, r in pmap[period].items():
            if r['pool_type'] not in SCURVE_POOL_TYPES: continue
            rm = r.get('removal_reason', '').strip()
            iv = rm == '1'; ii = rm in ('2','3','4','6')
            dis = nxt is not None and lid not in nl
            mdq = r.get('months_dq', 0) or 0
            if dis and not iv and not ii:
                iv = mdq == 0; ii = mdq > 0
            fp = pd_(r.get('first_pay_date','')); lm = pd_(r.get('loan_maturity_date',''))
            le = pd_(r.get('lockout_end_date','')); pe = pd_(r.get('prepay_end_date',''))
            age = mb(fp, od); rem = mb(od, lm)
            il = 1 if pd.notna(le) and od < le else 0
            ip = 1 if pd.notna(pe) and od < pe else 0
            pa = 1 if not il and not ip else 0
            lr = r.get('loan_rate', np.nan); sr = r.get('security_rate', np.nan)
            smm = np.nan
            if nxt and lid in nl:
                cu = r.get('upb', np.nan); nu = nl[lid].get('upb', np.nan)
                if pd.notna(cu) and pd.notna(nu) and cu > 0: smm = 1-(nu/cu)
            rows.append({
                'period': period, 'obs_date': od, 'loan_id': lid,
                'pool_cusip': r['pool_cusip'], 'pool_number': r['pool_number'],
                'pool_type': r['pool_type'], 'case_number': r.get('case_number',''),
                'agency_type': r.get('agency_type',''),
                'fha_program_code': r.get('fha_program_code',''),
                'loan_rate': lr, 'security_rate': sr,
                'servicing_spread': (lr-sr) if pd.notna(lr) and pd.notna(sr) else np.nan,
                'benchmark_rate': np.nan, 'refi_incentive_bps': np.nan,
                'orig_prin_bal': r.get('orig_prin_bal', np.nan),
                'upb_at_issuance': r.get('upb_at_issuance', np.nan),
                'current_upb': r.get('upb', np.nan),
                'pool_security_rpb': r.get('security_rpb', np.nan),
                'rpb_factor': r.get('rpb_factor', np.nan),
                'origination_date': r.get('origination_date',''),
                'first_pay_date': r.get('first_pay_date',''),
                'loan_maturity_date': r.get('loan_maturity_date',''),
                'loan_term_months': r.get('loan_term', 0),
                'age_months': age, 'remaining_term_months': rem,
                'lockout_term_yrs': r.get('lockout_term_yrs',0),
                'lockout_end_date': r.get('lockout_end_date',''),
                'prepay_premium_period_yrs': r.get('prepay_premium_period_yrs',0),
                'prepay_end_date': r.get('prepay_end_date',''),
                'prepay_penalty_flag': r.get('prepay_penalty_flag',''),
                'prepay_desc': r.get('prepay_desc',''),
                'in_lockout': il, 'in_prepay_penalty': ip, 'past_all_restrictions': pa,
                'months_to_lockout_end': max(0, mb(od,le) or 0) if pd.notna(le) and od<le else 0,
                'months_to_prepay_end': max(0, mb(od,pe) or 0) if pd.notna(pe) and od<pe else 0,
                'prepaid_voluntary': 1 if iv else 0,
                'prepaid_involuntary': 1 if ii else 0,
                'prepaid_any': 1 if iv or ii else 0,
                'removal_reason': rm, 'liquidation_flag': r.get('liquidation_flag',''),
                'months_dq': mdq, 'modified_ind': r.get('modified_ind',''),
                'insurance_type': r.get('insurance_type',''),
                'property_name': r.get('property_name',''),
                'property_state': r.get('property_state',''),
                'property_city': r.get('property_city',''),
                'msa': r.get('msa',''), 'num_units': r.get('num_units',0),
                'green_status': r.get('green_status',''),
                'affordable_status': r.get('affordable_status',''),
                'non_level_ind': r.get('non_level_ind',''),
                'mature_loan_flag': r.get('mature_loan_flag',''),
                'issuer_number': r.get('issuer_number',''),
                'issuer_name': r.get('issuer_name',''),
                'smm_approx': smm,
            })
    return pd.DataFrame(rows)


def build_summary(df):
    rows = []
    for period in sorted(df['period'].unique()):
        p = df[df['period']==period]; n = len(p)
        upb = p['current_upb'].sum(); vol = p['prepaid_voluntary'].sum()
        inv = p['prepaid_involuntary'].sum()
        wr = np.average(p['loan_rate'].dropna(),
            weights=p.loc[p['loan_rate'].notna(),'current_upb'].clip(lower=.01)
        ) if p['loan_rate'].notna().any() else np.nan
        wa = np.average(p['age_months'].dropna(),
            weights=p.loc[p['age_months'].notna(),'current_upb'].clip(lower=.01)
        ) if p['age_months'].notna().any() else np.nan
        rows.append({
            'Period': period, 'Loans': n, 'UPB ($M)': upb/1e6,
            'Vol': int(vol), 'Invol': int(inv),
            'CPR (ann %)': (1-(1-vol/max(n,1))**12)*100,
            'WA Rate': wr, 'WA Age': wa,
            '% Lock': p['in_lockout'].mean()*100,
            '% Pen': p['in_prepay_penalty'].mean()*100,
            '% Open': p['past_all_restrictions'].mean()*100,
        })
    return pd.DataFrame(rows)


def build_scurve_buckets(df):
    d = df[df['loan_rate'].notna()].copy()
    d['b'] = (d['loan_rate']*4).round()/4
    g = d.groupby('b').agg(n=('loan_id','count'), upb=('current_upb','sum'),
        pp=('prepaid_voluntary','sum'), age=('age_months','mean'),
        lk=('in_lockout','mean'), op=('past_all_restrictions','mean')).reset_index()
    g['smm'] = g['pp']/g['n'].clip(lower=1)
    g['cpr'] = (1-(1-g['smm'])**12)*100
    g['upb'] /= 1e6; g['lk'] *= 100; g['op'] *= 100
    return g.rename(columns={'b':'Rate (%)','n':'Loans','upb':'UPB ($M)',
        'pp':'Prepaid','age':'WA Age','lk':'% Locked','op':'% Open',
        'smm':'SMM','cpr':'CPR (ann %)'})[
        ['Rate (%)','Loans','UPB ($M)','Prepaid','SMM','CPR (ann %)','WA Age','% Locked','% Open']
    ].sort_values('Rate (%)')


def build_lockout(df):
    cats = [('In Lockout', df[df['in_lockout']==1]),
            ('In Penalty', df[(df['in_lockout']==0)&(df['in_prepay_penalty']==1)]),
            ('Open', df[df['past_all_restrictions']==1])]
    rows = []
    for lbl, sub in cats:
        n = len(sub)
        if n == 0: continue
        rows.append({
            'Category': lbl, 'Loans': n, 'UPB ($M)': sub['current_upb'].sum()/1e6,
            'Vol': int(sub['prepaid_voluntary'].sum()),
            'SMM': sub['prepaid_voluntary'].sum()/max(n,1),
            'CPR (ann %)': (1-(1-sub['prepaid_voluntary'].sum()/max(n,1))**12)*100,
            'WA Rate': np.average(sub['loan_rate'].dropna(),
                weights=sub.loc[sub['loan_rate'].notna(),'current_upb'].clip(lower=.01)
            ) if sub['loan_rate'].notna().any() else np.nan,
            'WA Age': sub['age_months'].mean(),
        })
    return pd.DataFrame(rows)


def fmt_wb(wb):
    hf = Font(name='Arial',bold=True,size=10,color='FFFFFF')
    hfill = PatternFill('solid',fgColor='1E3A5F')
    for ws in wb.worksheets:
        for c in ws[1]: c.font=hf; c.fill=hfill; c.alignment=Alignment(horizontal='center',wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row: c.font=Font(name='Arial',size=9); c.border=Border(bottom=Side('thin',color='D0D0D0'))
        for col in range(1,ws.max_column+1):
            mx = max((len(str(ws.cell(r,col).value or '')) for r in range(1,min(ws.max_row+1,50))),default=8)
            ws.column_dimensions[get_column_letter(col)].width = min(mx+3,25)
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions


def write_output(df, monthly_data):
    print(f"\nWriting {OUTPUT_FILE}...")
    pcols = ['period','loan_id','pool_cusip','pool_number','pool_type',
        'case_number','fha_program_code','agency_type',
        'loan_rate','security_rate','servicing_spread','benchmark_rate','refi_incentive_bps',
        'current_upb','orig_prin_bal','rpb_factor',
        'age_months','remaining_term_months','loan_term_months',
        'in_lockout','in_prepay_penalty','past_all_restrictions',
        'months_to_lockout_end','months_to_prepay_end','lockout_end_date','prepay_end_date','prepay_desc',
        'prepaid_voluntary','prepaid_involuntary','prepaid_any','removal_reason',
        'months_dq','modified_ind','property_state','num_units',
        'green_status','affordable_status','issuer_name','smm_approx']
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as w:
        build_summary(df).to_excel(w, sheet_name='Summary', index=False)
        build_scurve_buckets(df).to_excel(w, sheet_name='S-Curve Buckets', index=False)
        build_lockout(df).to_excel(w, sheet_name='Lockout Analysis', index=False)
        df[pcols].to_excel(w, sheet_name='Loan Panel', index=False)
        df.to_excel(w, sheet_name='Full Detail', index=False)
    wb = load_workbook(OUTPUT_FILE); fmt_wb(wb)
    ws = wb.create_sheet('Instructions',0)
    for r in [['GNMA MF Prepayment S-Curve Dataset'],[''],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['Periods:', ', '.join(sorted(monthly_data.keys()))],
        ['Observations:', len(df)], ['Unique Loans:', df['loan_id'].nunique()],
        ['Vol Prepays:', int(df['prepaid_voluntary'].sum())], [''],
        ['Fill benchmark_rate, then: refi_incentive_bps = (loan_rate - benchmark) * 10000']]:
        ws.append(r)
    ws['A1'].font = Font(name='Arial',bold=True,size=14,color='1E3A5F')
    ws.column_dimensions['A'].width=25; ws.column_dimensions['B'].width=60
    wb.save(OUTPUT_FILE)
    print(f"\n{'='*60}")
    print(f"  ok {OUTPUT_FILE}")
    print(f"     {len(df)} obs, {df['loan_id'].nunique()} loans, {int(df['prepaid_voluntary'].sum())} vol prepays")
    print(f"     Tabs: Instructions | Summary | S-Curve Buckets | Lockout | Loan Panel | Full Detail")
    print(f"{'='*60}")


# ═══════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="GNMA MF Downloader & S-Curve Builder")
    ap.add_argument("--email", help="GNMA email")
    ap.add_argument("--answer", help="Security answer")
    ap.add_argument("--months", type=int, default=6)
    ap.add_argument("--skip-download", action="store_true")
    ap.add_argument("--data-dir", help="Override data directory")
    ap.add_argument("--browser", action="store_true", help="Force Playwright")
    ap.add_argument("--headed", action="store_true", help="Show browser (debug)")
    args = ap.parse_args()

    global DATA_DIR, OUTPUT_FILE
    if args.data_dir:
        DATA_DIR = args.data_dir
        OUTPUT_FILE = os.path.join(args.data_dir, "gnma_mf_scurve_dataset.xlsx")

    print("""
+==============================================================+
|  GNMA Multifamily Data Downloader & S-Curve Builder          |
|  Auth: requests -> Playwright (auto-fallback)                |
|  Downloads mfplmon3, identifies prepayments, builds S-curve  |
+==============================================================+
""")

    if not args.skip_download:
        email = args.email or input("GNMA Disclosure email: ").strip()
        answer = args.answer or input("Security question answer: ").strip()
        success = False

        # Method 1: requests
        if not args.browser:
            dl = RequestsDownloader(email, answer)
            if dl.authenticate():
                download_all(dl, args.months)
                success = True

        # Method 2: Playwright (auto-fallback)
        if not success:
            if not args.browser:
                print("\n" + "-"*60)
                print("  Falling back to Playwright (headless browser)...")
                print("-"*60)
            if ensure_playwright():
                dl = PlaywrightDownloader(email, answer, headless=not args.headed)
                try:
                    if dl.authenticate():
                        download_all(dl, args.months)
                        success = True
                finally:
                    dl.close()

        if not success:
            print(f"""
+==============================================================+
|  Both auth methods failed.                                   |
|                                                              |
|  Manual fallback:                                            |
|  1. Go to ginniemae.gov disclosure data download page        |
|  2. Log in, download mfplmon3_YYYYMM.zip (6 months)         |
|  3. Place in: {DATA_DIR:<44} |
|  4. Re-run with: --skip-download                             |
|                                                              |
|  Debug tip: re-run with --headed to see the browser          |
+==============================================================+
""")

    # ─── PARSE ─────────────────────────────────────
    os.makedirs(DATA_DIR, exist_ok=True)
    allf = sorted(glob.glob(os.path.join(DATA_DIR,"*.zip")) + glob.glob(os.path.join(DATA_DIR,"*.txt")))
    if not allf:
        print(f"No files in {DATA_DIR}"); return

    print(f"\nParsing {len(allf)} files...")
    md = {}
    for fp in allf:
        m = re.search(r'(\d{6})', os.path.basename(fp))
        if not m: continue
        per = m.group(1)
        recs = read_mfplmon3(fp)
        if recs:
            md[per] = recs
            pt = defaultdict(int)
            for r in recs: pt[r['pool_type']] += 1
            print(f"  {os.path.basename(fp)} -> {per}: {len(recs)} records {dict(pt)}")

    if len(md) < 2:
        print("\nNeed 2+ months for prepay identification."); return

    df = build_panel(md)
    print(f"\nPanel: {len(df)} obs, {df['loan_id'].nunique()} loans")
    print(f"  Vol prepays: {df['prepaid_voluntary'].sum():.0f}")
    print(f"  Invol: {df['prepaid_involuntary'].sum():.0f}")
    write_output(df, md)


if __name__ == "__main__":
    main()
